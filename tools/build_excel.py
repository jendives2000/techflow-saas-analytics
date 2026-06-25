"""
Build project1_saas_analytics/dashboards/techflow_kpi.xlsx
3 sheets: RawData (Excel Table), KPIs (aggregates + chart data), Dashboard (KPI cells + 3 charts)
"""
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import os
from pathlib import Path

from dotenv import load_dotenv
import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

load_dotenv()

PG_HOST     = os.getenv("PG_HOST", "localhost")
PG_PORT     = int(os.getenv("PG_PORT", "5432"))
PG_USER     = os.getenv("PG_USER", "postgres")
PG_PASSWORD = os.getenv("PG_PASSWORD", "")
PG_DB       = "techflow"

OUT = Path("dashboards/techflow_kpi.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

PLAN_ORDER  = {"Month-to-month": 1, "One year": 2, "Two year": 3}
PLAN_LABELS = {"Month-to-month": "Starter", "One year": "Growth", "Two year": "Enterprise"}


def load_data() -> pd.DataFrame:
    conn = psycopg2.connect(
        host=PG_HOST, port=PG_PORT, user=PG_USER,
        password=PG_PASSWORD, database=PG_DB
    )
    df = pd.read_sql("""
        SELECT customer_id, plan_type, subscription_months,
               monthly_mrr, lifetime_value, churned, payment_method,
               CASE
                 WHEN subscription_months < 6  THEN 'New (0-6mo)'
                 WHEN subscription_months < 24 THEN 'Growing (6-24mo)'
                 ELSE 'Mature (24mo+)'
               END AS lifecycle_stage
        FROM customers
        ORDER BY customer_id
    """, conn)
    conn.close()
    return df


def write_rawdata(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    last_col = get_column_letter(len(df.columns))
    tab = Table(displayName="CustomerData", ref=f"A1:{last_col}{len(df) + 1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)


def write_kpis(ws, df: pd.DataFrame) -> dict:
    """Write KPI summary (rows 1-7). Returns {label: row_number} for Dashboard references."""
    active = df[df["churned"] == "No"]
    kpis = [
        ("Total MRR (Active)", active["monthly_mrr"].sum(),                                   "$#,##0"),
        ("Churn Rate %",       (df["churned"] == "Yes").sum() / len(df),                      "0.00%"),
        ("Avg LTV",            df["lifetime_value"].mean(),                                    "$#,##0"),
        ("Avg Tenure Months",  active["subscription_months"].mean(),                           "0.0"),
        ("At-Risk MRR",        active[active["subscription_months"] < 6]["monthly_mrr"].sum(), "$#,##0"),
        ("LTV MRR Ratio",      df["lifetime_value"].mean() / df["monthly_mrr"].mean(),         "0.0"),
    ]
    ws["A1"].value = "Metric"
    ws["A1"].font  = Font(bold=True)
    ws["B1"].value = "Value"
    ws["B1"].font  = Font(bold=True)

    row_map = {}
    for i, (label, value, fmt) in enumerate(kpis, start=2):
        ws.cell(row=i, column=1).value = label
        c = ws.cell(row=i, column=2)
        c.value         = value
        c.number_format = fmt
        row_map[label]  = i
    return row_map


def write_chart_tables(ws, df: pd.DataFrame) -> tuple:
    """Write 3 data tables for charts. Returns (MRR_ROW, CHURN_ROW, LTV_ROW)."""
    active = df[df["churned"] == "No"]

    # MRR by Plan (start row 10)
    MRR_ROW = 10
    ws.cell(row=MRR_ROW, column=1).value = "Plan"
    ws.cell(row=MRR_ROW, column=1).font  = Font(bold=True)
    ws.cell(row=MRR_ROW, column=2).value = "Total MRR"
    ws.cell(row=MRR_ROW, column=2).font  = Font(bold=True)
    mrr = (active.groupby("plan_type")["monthly_mrr"].sum()
           .reset_index()
           .assign(order=lambda d: d["plan_type"].map(PLAN_ORDER))
           .sort_values("order"))
    for i, (_, row) in enumerate(mrr.iterrows(), start=MRR_ROW + 1):
        ws.cell(row=i, column=1).value = PLAN_LABELS[row["plan_type"]]
        c = ws.cell(row=i, column=2)
        c.value         = row["monthly_mrr"]
        c.number_format = "$#,##0"

    # Churn Rate by Plan (start row 16)
    CHURN_ROW = 16
    ws.cell(row=CHURN_ROW, column=1).value = "Plan"
    ws.cell(row=CHURN_ROW, column=1).font  = Font(bold=True)
    ws.cell(row=CHURN_ROW, column=2).value = "Churn Rate"
    ws.cell(row=CHURN_ROW, column=2).font  = Font(bold=True)
    churn = (df.groupby("plan_type")
               .agg(churn_rate=("churned", lambda x: (x == "Yes").sum() / len(x)))
               .reset_index()
               .assign(order=lambda d: d["plan_type"].map(PLAN_ORDER))
               .sort_values("order"))
    for i, (_, row) in enumerate(churn.iterrows(), start=CHURN_ROW + 1):
        ws.cell(row=i, column=1).value = PLAN_LABELS[row["plan_type"]]
        c = ws.cell(row=i, column=2)
        c.value         = row["churn_rate"]
        c.number_format = "0.00%"

    # LTV Distribution (start row 22)
    LTV_ROW = 22
    ws.cell(row=LTV_ROW, column=1).value = "LTV Range"
    ws.cell(row=LTV_ROW, column=1).font  = Font(bold=True)
    ws.cell(row=LTV_ROW, column=2).value = "Customers"
    ws.cell(row=LTV_ROW, column=2).font  = Font(bold=True)
    bins   = [0, 500, 1000, 2000, 3000, 5000, 10000]
    labels = ["$0-500", "$500-1K", "$1K-2K", "$2K-3K", "$3K-5K", "$5K+"]
    ltv_series = df["lifetime_value"].dropna()
    counts = (pd.cut(ltv_series, bins=bins, labels=labels, include_lowest=True)
               .value_counts()
               .reindex(labels))
    for i, (label, count) in enumerate(counts.items(), start=LTV_ROW + 1):
        ws.cell(row=i, column=1).value = label
        ws.cell(row=i, column=2).value = int(count)

    return MRR_ROW, CHURN_ROW, LTV_ROW


def make_bar_chart(title: str, y_title: str, data_ref, cat_ref, y_fmt: str = None) -> BarChart:
    chart           = BarChart()
    chart.type      = "col"
    chart.title     = title
    chart.y_axis.title = y_title
    chart.style     = 10
    chart.width     = 15
    chart.height    = 10
    if y_fmt:
        chart.y_axis.numFmt = y_fmt
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    return chart


def write_dashboard(ws_dash, ws_kpi, row_map: dict,
                    MRR_ROW: int, CHURN_ROW: int, LTV_ROW: int) -> None:
    ws_dash["A1"]      = "TechFlow SaaS — Executive Dashboard"
    ws_dash["A1"].font = Font(bold=True, size=16)

    kpi_layout = [
        ("A3", "B3", "Total MRR",    "Total MRR (Active)", "$#,##0"),
        ("D3", "E3", "Churn Rate %", "Churn Rate %",       "0.00%"),
        ("A5", "B5", "Avg LTV",      "Avg LTV",            "$#,##0"),
        ("D5", "E5", "Avg Tenure",   "Avg Tenure Months",  '0.0" mo"'),
    ]
    for lbl_cell, val_cell, display, key, fmt in kpi_layout:
        ws_dash[lbl_cell].value = display
        ws_dash[lbl_cell].font  = Font(bold=True, size=11)
        ws_dash[val_cell].value = f"=KPIs!B{row_map[key]}"
        ws_dash[val_cell].number_format = fmt
        ws_dash[val_cell].font  = Font(bold=True, size=14)

    ws_dash.add_chart(make_bar_chart(
        "MRR by Plan Type", "MRR ($)",
        Reference(ws_kpi, min_col=2, min_row=MRR_ROW,     max_row=MRR_ROW   + 3),
        Reference(ws_kpi, min_col=1, min_row=MRR_ROW + 1, max_row=MRR_ROW   + 3),
    ), "A8")

    ws_dash.add_chart(make_bar_chart(
        "Churn Rate by Plan Type", "Churn Rate",
        Reference(ws_kpi, min_col=2, min_row=CHURN_ROW,     max_row=CHURN_ROW + 3),
        Reference(ws_kpi, min_col=1, min_row=CHURN_ROW + 1, max_row=CHURN_ROW + 3),
        y_fmt="0%"
    ), "I8")

    ws_dash.add_chart(make_bar_chart(
        "LTV Distribution", "Customers",
        Reference(ws_kpi, min_col=2, min_row=LTV_ROW,     max_row=LTV_ROW + 6),
        Reference(ws_kpi, min_col=1, min_row=LTV_ROW + 1, max_row=LTV_ROW + 6),
    ), "A28")


def main():
    print("Loading data from PostgreSQL...")
    df = load_data()
    print(f"  {len(df):,} rows loaded")

    wb     = Workbook()
    ws_raw = wb.active
    ws_raw.title = "RawData"
    write_rawdata(ws_raw, df)
    print("  RawData: done")

    ws_kpi = wb.create_sheet("KPIs")
    row_map = write_kpis(ws_kpi, df)
    MRR_ROW, CHURN_ROW, LTV_ROW = write_chart_tables(ws_kpi, df)
    print("  KPIs: done")

    ws_dash = wb.create_sheet("Dashboard")
    write_dashboard(ws_dash, ws_kpi, row_map, MRR_ROW, CHURN_ROW, LTV_ROW)
    print("  Dashboard: done")

    wb.save(OUT)
    print(f"\nSaved → {OUT}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
